# print('mac keveb')
# price = 10
# price2 = price + 60
# rating = 4.9
# name = 'mac'
# is_published = True
# print(price2)
# print('o----')
# print(' |||||')
# print('*' * 8)
# full_name = 'John Smith'
# age = 20
# is_new = True
# print()
# # print('hello ' + name)
# namer = input('what is your name? ')
# color = input('what is your favorite colour? ')
# print (namer + ' likes ' + color)
# birth_year = input('Birth year: ')
# age = 2022 - int(birth_year)
# print(age)
# wek = input('what is your weight? (lbs)): ')
# ew = int(wek) * 0.45
# print(ew)
# core = "python's core for pro"
# course2 = 'python for "pros'
# # for email
# EmailMessage = '''
# Hi Mc,

# Here is our first email to you.

# Thank you,
# The Technical Team
# '''
# print(EmailMessage)

# # looking at other characteristic of strings in python
# # from __future__ import division
# # from tokenize import Exponent
# # from turtle import xcor


# curve = 'python for pros'
# print(curve[1:-2])
# new_curve = curve[:]
# print(new_curve)
# plus = 'Jennifer'
# print(plus[1:-1])
# # formatted strings
# first = 'John'
# last = 'Smith'
# message = first + ' [' + last  + '] is a coder'
# msg = f'{first} [{last}] is a coder'
# # John [Smith] is a Coder
# print(msg)
# print(message)
# # strings method
# meme = 'python for pros'
# # to count characters
# print(len(meme))
# # to capitalize
# print(meme.upper())
# # for lowercase
# print(meme.lower())
# # to find the index or sequence of characters
# print(meme.find('o'))
# # to replace an item(in a string)
# print(meme.replace('pros', 'experts'))
# # to confirm an item(booleans)
# print('python' in meme)
# # arithmetic operations
# print(10 + 3)
# print(10 // 3)
# print(10 ** 3)
# x = 3
# x -= 33
# print(x)
# # operation precedence
# x1 = 10 + 3 * 2 ** 4
# ki = 3 * 2 ** 4 + 10
# print(ki)
# print(x1)
# # parenthesis
# # Exponentiation = 2 ** 3
# # multiplication or division
# # addition or subtraction
# # print(Exponentiation)

# # maths function
# x = 2.9
# print(x)
# print(round(x))
# print(abs(-2.9))
# import math

# print(math.ceil(2.9))
# # the floor method
# print(math.floor(2.1))
# #  check out for python 3 math module
# # if statements
# is_hot = False
# is_cold = False
# if is_hot:
#     print("it's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("it's a cold day")
#     print("wear a warm clothes")
# else:
#    print("it's a lovely day")
# print("Enjoy your day")

# price = 1000000
# has_good_credit = True
# if has_good_credit:
#     down_payment = 0.1 * price
# else:
#     down_payment = 0.2 * price
# print(f"DOWN PAYMENT: ${down_payment}")
# # # logical operators
# has_high_income = False
# has_good_credit = True
# has_criminal_record = True

# if has_high_income and has_good_credit:
#     print("Eligible for loan")
# else:
#     print("sorry, you are currently not Eligible for this loan")

# if has_high_income or has_good_credit and not has_criminal_record:
#     print("Eligible for loan")
# else:
#     print("sorry, you are currently not Eligible for this loan")

# tempo = input('what is your tempo range? ')

# if tempo != 35:
#     print("it's a hot day")
# else:
#     print("it's npt a hpt day")

# name = "t"
# if len(name) < 3:
#     print("name must be at least 3 characters")
# elif len(name) > 50:
#     print("name must be a maximum of 50 characters")
# else:
#     print("name looks good!")


# formatted by myself
# name = input('what is your name? ')
# if len(name) < 3:
#     print("name must be at least 3 characters")
# elif len(name) > 50:
#     print("name must be a maximum of 50 characters")
# else:
#     print("hello there!")
#     print("name looks good!")
# # weight conversion code vip
# weight = int(input('weight: '))
# unit = input('(L)bs or (K)g: ')
# if unit.upper == "L":
#     converted = weight * 0.45
#     print(f"You are {converted} kilos")
# else:
#     converted = weight // 0.45
#     print(f"You are {converted} pounds")
# # make sure to do same but in this case, for height and distance

#  now while loops
# i = 1
# while i <= 5:
#     print('*' * i)
#     i = i + 1
# print("done!")

# # guessing game
# secret_number = 9
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit:
#     guess = int(input('guess: '))
#     guess_count += 1
#     if guess == secret_number:
#         print('You Won!')
#         break
# else:
#     print("Please, try again later")
# # car game
# # multiple lines of repeated .lower
# command = ""
# while command.lower() != "quit":
#     command = input("> ")
#     if command.lower() == "start":
#         print("car started...Ready to go")
#     elif command.lower() == "stop":
#         print("car stopped.")
#     elif command.lower() == "help":
#         print("""
# start - to start the car
# stop - to stop the car
# quit - to exit""")
#     elif command.lower() == "quit":
#         break
#     else:
#         print("sorry, type 'help' for support")

# # writing to=he above codes in a more professional way
# command = ""
# while True:
#     command = input("> ").lower()
#     if command == "start":
#         print("car started...Ready to go")

#     elif command == "stop":
#         print("car stopped.")

#     elif command == "help":
#         print("""
# start - to start the car
# stop - to stop the car
# quit - to exit""")
#     elif command == "quit":
#         break
#     else:
#         print("sorry, type 'help' for support")

# command = ""
# started = False
# while True:
#     command = input("> ").lower()
#     if command == "start":
#         if started:
#             print("car is already started!")
#         else:
#             started = True
#             print("car started...")
#     elif command == "stop":
#         if not started:
#             print("car is already stopped!")
#         else:
#             started = False
#             print("car stopped")
#     elif command == "help":
#         print("""
# start - to start the car
# stop - to stop the car
# quit - to exit""")
#     elif command == "quit":
#         break
#     else:
#         print("sorry, type 'help' for support")

# from dataclasses import field
# import numbers
# from tkinter import Y


# for item in 'python':
#     print(item)
# # for names
# for item in ['james', 'josh', 'mosh']:
#     print(item)
# # for numbers
# for i in ['1', '1', '4']:
#     print(i)
# # code
# power = input('type in a text in the field')
# for item in power:
#     print(item)
# # using the range function
# for item in range(5, 10, 2):
#     print(item)
# prices = [10, 20, 30]

# total = 0
# for price in prices:
#     total += price
# print(f"Total: {total}")

# # nested loops
# # (x, y)
# # (0, 0)
# # (0, 1)
# # (0, 2)
# # (1, 0)
# # (1, 1)
# # (1, 2)
# for x in range(4):
#     for y in range(3):
#         print(f'({x}, {y})')
# numbers = [6,2,6,2,6]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)


# numbers = [2,2,2,6]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)


# numbers = [2,2,2,2]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)

# names = ['john', 'bob', 'mosh', 'sarah', 'mary']
# print(names[1])
# print(names[2])
# print(names[3])
# print(names[-2:])
# names[0] = 'raysz'
# print(names)

# # number = [3, 5, 6, 8, 10, 100]
# # print(max(number))

# number = [3, 5, 6, 8, 10, 100]
# max = number[0]
# for number in number:
#     if number > max:
#         max = number
# print(max)
# 2 dimensional list in python
# matrix = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9]
# ]
# # matrix[0][3] = 20
# # print(matrix[0][2])
# for row in matrix:
#     for item in row:
#         print(item)

# # list method
# numbers = [5, 2, 1, 5, 7, 5, 4]
# numbers.append(20)
# print(numbers)

# numbers.insert(1, 67)
# print(numbers)

# numbers.remove(2)
# print(numbers)

# # numbers.clear()
# # print(numbers)

# numbers.pop()
# print(numbers)
# print(50 in numbers)

# names = "hello world"
# print('james' in names)
# print('w' in names)

# print(numbers.count(5))
# print(numbers.sort())

# numbers.sort()
# print(numbers)

# numbers.reverse()
# print(numbers)

# numbers2 = numbers.copy()
# print(numbers2)
# numbers2.sort()
# print(numbers2)
# #removing duplicates on a list
# from unicodedata import numeric


# numbers = [2,3,2, 2, 2, 4, 5, 76, 9, 0]
# unique = []
# for number in numbers:
#     if number not in unique:
#         unique.append(number)
# print(unique)
# # tuples
# number = (1, 2, 3)
# print(number[0])
# # unpacking
# coordinates = (1, 2, 3)
# x = coordinates[0]
# y = coordinates[1]


# number = [1, 2, 3]
# print(number[0])
# # unpacking
# coordinates = (1, 2, 3)
# x = coordinates[0]
# y = coordinates[1]
# # instead of writing this long line of code by coding the following
# x, y, z = coordinates
# print(x)
# print(z)
# print(y)
# # dictionary
# customer = {
#     "name": "john smith",
#     "age": 30,
#     "is_verified": True
# }
# # to add new key value pairs to a dictionary, here we go!
# customer["DOB"] = "Dec, 1 1869"
# print(customer["DOB"], 'is your date of birth')

# cust = {
#     "name":input("what is your name? "),
#     "age":input("how old are you? "),
#     "is_verified": True
# }
# print(cust)

# # numbers to word translator
# Phone = input("Phone: ")
# digits_mapping = {
#     "0": "Zero",
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four",
#     "5": "Five",
#     "6": "Six",
#     "7": "Seven",
#     "8": "Eight",
#     "9": "Nine",
#     "+": "Plus"
# }
# output = ""
# for character in Phone:
#     output += digits_mapping.get(character, "!") + " "
# print(output)


# Amount = input("Amount: ")
# digits_mapping = {
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four",
#     "5": "Five",
#     "6": "Six",
#     "7": "Seven",
#     "8": "Eight",
#     "9": "Nine",
#     "00": "Hundred",
#     "000": "Thousand",
#     "000000": "Million",
#     "000000000": "Billion",
#     "000000000000": "Trillion"
# }
# Amount = ""
# for character in Amount:
#     Amount += digits_mapping.get(character, "!") + " "
# print(Amount)
# # go back to time 2:29:17
# message = input(">")
# words = message.split(' ')
# print(words)

# # functions
# def greet_user(first_name, last_name):
#     print(f'hello {first_name} {last_name}!')
#     print('welcome aboard')

# print("start")
# greet_user("bte", "poli")
# greet_user("grey", "hector")
# greet_user("inkiki", "clark")
# print("finish")

# # parameters these are placeholders of a defined in a function for receiving information
# # argument are the pieces of information supplied or received

# # keyword argument
# # return values

# def square(number):
#     return number * number

# print(square(3))

# def emoji_converter(message):
#     words = message.split(" ")
#     emojis= {
#         "H": "happy",
#         "g": "makeups"
#     }
#     output = ""
#     for l in words:
#         output += emojis.get(l, l) + " "
#     return output

# message = input(">")
# # handling errors in python
# try:
#     print(emoji_converter(message))
#     age = int(input('Age: '))
#     income = input('hourly income: ')
#     rist = income / age
#     print(age)
# except ZeroDivisionError:
#     print('Age cannot be 0.')
# except ValueError:
#     print('invalid error')

# classes in python
# classes define the blueprint or templates for creating an object while objects are instances based on that blueprint
# class Point:
#     def nove(self):
#         print("move")

#     def draw(self):
#         print("draw")


# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.x)
# point1.draw()

# point2 = Point()
# point2.x = 1
# print(point2.x)

# # constructor
# # a constructor is a function that get called at the time of creating an object
# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#     def move(self):
#         print("move")

#     def draw(self):
#         print("draw")


# point = Point(10, 20)
# point.x = 78
# print(point.x)

# class Person:
#     # def __init__(self, name, talk):
#     #     self.name = input('what is your name? ')
#     #     self.talk = input('wow us with your pitch! ')
#     def __init__(self, name):
#             self.name = name
#     def name(self):
#         print("name")

#     def talk(self):
#         print(f"Hi, i am {self.name}")
# # Person.talk("")
# # Person.name("")

# john = Person("John smith")
# # print(john.name)
# john.talk()
# # person = Person('james', 'goodness')
# bob = Person("VTU PLUS")
# bob.talk()

# inheritance in python
#  DRY DON'T REPEAT YOURSELF

# # parent class
#  Mammal:
#     def walk(self):
#         print("walk")


#  # child class which inherit functions from parents class
#  Dog(Mammal):
#     pass

# class Cat(Mammal):
#     pass

# dog1 = Dog()
# dog1.walk()
# cat1 = Cat()
# cat1.walk()

# # to add more attributes related to a specific item
# class Dog(Mammal):
#     def bark(self):
#         print("bark")

# class Cat(Mammal):
#     def smilie(self):
#         print("smilie")

# cat1 = Cat()
# cat1.smilie()

# modules in python
# we use modules to organise our codes just like sections in the super market

# def kg_to_lbs(weight):
#     return weight / 0.45

# def lbs_to_kg(weight):
#     return weight * 0.45

# import converters
# from converters import kg_to_lbs


# print(converters.kg_to_lbs(200))
# print(converters.lbs_to_kg(444))

# import utilities
# from utilities import find_max
# numbers = [50, 40, 10, 790]
# max = find_max(numbers)
# print(max)
# # packages
# #  packages are another of organizing our code
# # a package is a container for multiple modules
# # in file system term, it is a directory or a folder
# import ecommerce.shipping
# ecommerce.shipping.calc_shipping()
# #  to make the code look simple
# from ecommerce.shipping import calc_shipping

# calc_shipping()
# calc_shipping()
# calc_shipping()
# calc_shipping()

# # to access all functions inside a module in a package 
# from ecommerce import shipping

# shipping.calc_shipping()
# # generating random values
# import random

# for i in range(3):
#     print(random.random())
# # to generate random values within a specific range
# for m in range(3):
#     print(random.randint(10, 20))
# # making a random selection
# members = ['james', 'martha', 'bob', 'classic']
# leader = random.choice(members)
# print(leader)

# # rolling a dice
# import random


# class Dice:
#     def roll (self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         return first, second


# dice = Dice()
# print(dice.roll())
# # files and directory
# from pathlib import Path

# path = Path("ecommerce")
# print(path.exists())

# # to add a directory to a package
# path = Path("emails")
# # the below stands for MAKE DIRECTORY
# print(path.mkdir())
# to search for files in a directory
# path = Path()
# for file in path.glob('*'):
#     print(file)    
# Py-pip and Pip
# Excel spreadsheet processing
import openpyxl as xl
from openpyxl.chart import Reference, BarChart

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
# cell = sheet.cell(1, 1)
print(cell.value)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   min_col=4,
                   max_row=sheet.max_row,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_data(chart, 'e2')

wb.save('transactions2.xlsx')
