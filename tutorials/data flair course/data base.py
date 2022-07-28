# def students(**kwargs: object) -> object:
# def james(funds=None):
#     kind = input("Type in your name to see your examination score: ")
#     name = funds
#     score = 70
#     grade = "A"
#     if kind == name:
#         return f"Congratulations {name}.\nYou scored {score} on your spanish test with Grade {grade}"
#     else:
#         return f"Hey! {kind}\nName missing\nCheck the general office for documentation"
#
#
# print(james(funds="clak"))

# def femi(me=None):
#     kin = input("Type in your name to see your examination score: ")
#     nam = me
#     scor = '{0:8} | {1:9}'.format('score', 'Grade')
#     fr = '{0:8} | {1:9}'.format(70, 'A')
#     # grad = "B"
#     if kin == nam:
#         return f"Congratulations {nam}.\n{scor}\n{fr}\nsee your result on your spanish test above"
#     else:
#         return f"Hey! {kin}\nName missing\nCheck the general office for documentation"
#
# print(femi(me="cv"))


# for key, value in **kwargs.items():
#     print("%s || %s" % (key, value))
#
# time = students(name=input('Type in your name to see your examination score: '), score='75', Grade='B')
# name = "%s" % (students())
#
# sur_name = input("SurName: ")
# first_name = input("first name: ")
# #
# name = "%s, %s" % (sur_name, first_name)
#
#
# def score(num):
#     x = num
#     if x >= 70:
#         return f"Congratulations {name}.\nYou scored {x} on your spanish test"
#     else:
#         return f"Hey! {name}.\nyour score didn't meet the minimum requirement for the test.\nYou might want to sit for the next session examinations "
#
#
# # # Reading an excel file using Python
# # # import xlrd
# # import openpyxl as xl
# # # from openpyxl.chart import BarChart
# # loc = "transactionss.xlsx"
# #
# # # To open Workbook
# # wb = xl.load_workbook('transactions.xlsx')
# # sheet = wb.sheet_by_index(0)
# #
# # # For row 0 and column 0
# # print(sheet.cell_value(5))
#
# # import csv
# #
# # filename = "student_data.csv"
# # header = ("Rank", "Rating", "Title", "Grade")
# # data = [
# # (1, 9.2, "The shawshank redemption(1995)", "A"),
# # (2, 9.2, "The godfather redeem(1997)", "B"),
# # (3, 9., "The godfather re II(1990)", "C"),
# # (4, 8.9, "The godfather(1900)", "A")
# # ]
# #
# # def writer(header, data, filename):
# #     pass
# #
# # def writer(header, data, filename):
# #     with open(filename, "w", newline = "") as csvfile:
# #         movies = csv.writer(csvfile)
# #         movies.writerow(header)
# #         for x in data:
# #             movies.writerow(x)
#
# # Program to extract number of
# # columns in Python
# # import xl
# # import openpyxl as xl
# #
# # loc = ("transactionss.xlsx")
# #
# # wb = xl.load_workbook(loc)
# # sheet = wb.sheet_by_index(0)
# #
# # # For row 0 and column 0
# # sheet.cell_value(0, 1)
# #
# # # Extracting number of columns
# # print(sheet.ncols)
#
# from openpyxl.workbook import workbook
# from openpyxl import load_workbook
# from tabulate import tabulate
# # load existing spreedsheet
# wb = load_workbook('students docs.xlsx')
# # create an active worksheet
#
# ws = wb.active
#
# # set a variable
# name = ws["A3"].value
# color = ws["B3"].value
#
# # print something from our spreedsheet
# # print(f'{name}: {color}')
#
# # grab a whole column
# # column_a = ws['7']
#
# for loop
# for cell in column_a:
#     print(cell.value)
# print(cell.value)
# from tabulate import tabulate

# range = ws['B3':'D11']
# print("{:<8} {:<15} {:<10}".format(ws['B1':'C1':'D1']))
# print(range)
# print(pi)

# # loop
# for cell in range:
#     for x in cell:
#         print(x.value)

# from tabulate import tabulate
# import pandas as pd
#
# df = pd.read_file (r'students docs.xlsx')
# print(df)

#
# import os
# os.getcwd()
# # os.chdir()
#
# import pandas as pd
# file = 'students docs.xlsx'
# data = pd.ExcelFile(file)
# print(data.sheet1)
#
# df = data.parse('students')
# var = df.info
#
# print(df.head(10))
from tabulate import tabulate

d = [
    ["james cla", 70, "A"],
    ["kasty", 80, "A"],
    ["femi", 65, "B"],
    ["sent", 50, "C"],
    ["coady", 85, "A"],
    ["kelly", 60, "B"],
    ["camrolls", 40, "D"],
    ["kat", 82, "A"],
    ["fem", 55, "D"],
    ["sen", 50, "C"],
    ["coat", 85, "A"],
    ["key", 60, "B"]]

print("{:<12} {:<15} {:<10}".format('student_name', 'score', 'grade'))

for v in d:
    student_name, score, grade = v
    print("{:<12} {:<15} {:<10}".format(student_name, score, grade))

from openpyxl import workbook, load_workbook

wb = load_workbook('student doc.xlsx')
ws = wb.active
ws.active_cell
# ws['B3'].value = "tod"
# ws['A3'].value = 2
ws = wb.save('student doc.xlsx')

perse = [
    ["james cla", 70, "A"],
    ["kasty", 80, "A"],
    ["femi", 65, "B"],
    ["sent", 50, "C"],
    ["coady", 85, "A"],
    ["kelly", 60, "B"],
    ["camrolls", 40, "D"],
    ["kat", 82, "A"],
    ["fem", 55, "D"],
    ["sen", 50, "C"],
    ["coat", 85, "A"],
    ["key", 60, "B"]]


def person(perse):
    for i in perse:
        for v in i:
            print(i)


print(tabulate(perse))

james = ('james', 70, 'A'),
kasty = "kasty", 80, "A",
femi = "femi", 65, "B",
sent = "sent", 50, "C",
coady = "coady", 85, "A",
kelly = "kelly", 60, "B",
camroll = "camrolls", 40, "D",
kat = "kat", 82, "A",
fem = "fem", 55, "D",
sen = "sen", 50, "C",
coat = "coat", 85, "A",
key = "key", 60, "B"

print("{:<12} {:<15} {:<10}".format('student_name', 'score', 'grade'))
print(tabulate(james))