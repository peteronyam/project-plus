# from openpyxl import workbook, load_workbook, Workbook
#
# # wb = load_workbook('students docs.xlsx')
# # ws = wb.active
# # print(ws['A3'].value)
# # # to edit a value inside a cell
# # ws['A2'].value = 1
# # ws['B1'].value = "S.NAMES"
# # # wb.save('students docs.xlsx')
# # # accessing a specific sheet
# # ws = wb['students docs']
# # print(wb['students docs'])
# # # to add sheets to an excel file
# # wb.create_sheet("assignment")
# # print(wb.sheetnames)
# # creating a new workbook
# from openpyxl import workbook
# wb = Workbook()
# ws = wb.active
# ws.title = "Time"
#
# ws.append(['tin', 'is', 'kind'])
# ws.append(['tin', 'is', 'kind'])
# ws.append(['tin', 'is'])
# ws.append(['tin', 'is', 'kind'])
# ws.append(['tin', 'is', 'kind', 'tin', 'is', 'kind'])
# ws.append(['tin', 'is', 'kind'])
# ws.append(['tin', 'is', 'kind', 'tin', 'is', 'kind'])
#
# wb.save('Time.xlsx')

# # looping through a cell
# import openpyxl.styles.fills
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
#
# # from openpyxl.utils import get_column_letter
# # from tabulate import tabulate
#
# wb = load_workbook('students doc.xlsx')
# ws = wb.active
# # for row in range(1,11):
# #     for col in range(1, 4):
# #         # char = chr(65 + col)
# #         char = get_column_letter(col)
# #         ws[char + str(row)] = char + str(row)
# # to insert rows and col
# ws.insert_rows(3)
# ws['B3'].value = "teams coad"
# ws['C3'].value = 68
# ws['D3'].value = "B"
# ws.move_range("B1:C5", rows=3, cols=3)
# wb.save('student doc.xlsx')
import openpyxl.worksheet.table
from openpyxl import workbook, load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


data = {
    "james": {
        "math": 65,
        "science": 78,
        "english": 68,
        "art": 88
    },
    "kent": {
            "math": 68,
            "science": 70,
            "english": 60,
            "art": 80
        },
    "Tim": {
            "math": 80,
            "science": 76,
            "english": 66,
            "art": 82
        },
    "mas": {
            "math": 65,
            "science": 78,
            "english": 68,
            "art": 88
        },
    "sam": {
            "math": 65,
            "science": 78,
            "english": 68,
            "art": 88
        },
    "jane": {
            "math": 40,
            "science": 60,
            "english": 54,
            "art": 100
        }
}

# wb = Workbook()
# ws = wb.active
# ws.title = "Grades"
#
# headings = ['Name'] + list(data['james'].keys())
# ws.append(headings)
#
# for person in data:
#     grades = list(data[person].values())
#     ws.append([person] + grades)
#
# for col in range(2, len(data['james']) + 2):
#     char = get_column_letter(col)
#     ws[char + "8"] = f"=SUM({char + '2'}:{char + '7'})/{len(data)}"
# for col in range(1, 6):
#     ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")
#
# wb.save("NewGrades.xlsx")